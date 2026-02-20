"""
convert_excel_to_json.py
Convierte el Excel de ganadería a un JSON compatible con la app PWA offline.
Uso:
  python convert_excel_to_json.py "TuArchivo.xlsx"

Salida:
  ganaderia_import.json
"""
import sys, re, json, datetime
from pathlib import Path
import openpyxl


def norm(s):
    if s is None:
        return ""
    return str(s).strip()


def parse_date(v):
    if v is None or v == "":
        return ""
    if isinstance(v, (datetime.date, datetime.datetime)):
        return v.date().isoformat() if isinstance(v, datetime.datetime) else v.isoformat()
    s = str(v).strip()
    m = re.match(r"^\d{4}-\d{2}-\d{2}$", s)
    if m:
        return s
    m = re.search(r"(\d{1,2})[-/](\d{1,2})[-/](\d{2,4})", s)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if y < 100:
            y += 2000
        try:
            return datetime.date(y, mo, d).isoformat()
        except ValueError:
            return ""
    return ""


def uid(prefix, i):
    return f"{prefix}_import_{i}"


def find_header_map(row):
    mp = {}
    for idx, v in enumerate(row):
        key = norm(v).lower()
        if key:
            mp[key] = idx
    return mp


def safe_get(row, mp, *names):
    for n in names:
        i = mp.get(n.lower())
        if i is not None and i < len(row):
            return row[i]
    return None


def slug_key(s: str) -> str:
    s = norm(s).lower()
    s = re.sub(r"\s+", "_", s)
    s = re.sub(r"[^a-z0-9_áéíóúñ]", "", s)
    return s.strip("_")


def main():
    if len(sys.argv) < 2:
        print('Uso: python convert_excel_to_json.py "archivo.xlsx"')
        sys.exit(1)

    xlsx = Path(sys.argv[1]).expanduser().resolve()
    if not xlsx.exists():
        print(f"No existe: {xlsx}")
        sys.exit(2)

    wb = openpyxl.load_workbook(xlsx, data_only=True)

    # ✅ Incluye brutos y meds
    data = {k: [] for k in [
        "users", "animals", "brutos", "meds",
        "milk", "healthEvents", "boosters", "repro",
        "salesCheese", "buyMilk", "transMilk", "fixedCosts"
    ]}

    # Default user
    data["users"].append({"id": "user_import", "name": "Importado"})

    # --- Animals from Bovinos_Activos (guardando extras)
    if "Bovinos_Activos" in wb.sheetnames:
        ws = wb["Bovinos_Activos"]
        header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        mp = find_header_map(header)
        idx = 0

        for r in range(2, ws.max_row + 1):
            row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]

            arete = safe_get(row, mp, "arete")
            nombre = safe_get(row, mp, "nombre/arete", "nombre")
            finca = safe_get(row, mp, "finca")
            sexo = safe_get(row, mp, "sexo")
            raza = safe_get(row, mp, "raza")

            if arete is None and nombre is None:
                continue

            arete_s = norm(arete)
            if not arete_s and re.fullmatch(r"\d+", norm(nombre)):
                arete_s = norm(nombre)
                nombre = ""

            if not arete_s and not norm(nombre):
                continue

            extras = {}
            for i, h in enumerate(header):
                key = slug_key(h)
                if not key:
                    continue
                val = row[i] if i < len(row) else None
                if val is None or str(val).strip() == "":
                    continue
                extras[key] = val

            idx += 1
            data["animals"].append({
                "id": uid("ani", idx),
                "arete": arete_s,
                "name": norm(nombre),
                "finca": norm(finca),
                "sexo": norm(sexo) or "Hembra",
                "raza": norm(raza),
                "extras": extras,
                "createdBy": "user_import",
                "createdAt": 0
            })

    # Helper: map arete->animalId and name->animalId
    by_arete = {a["arete"]: a["id"] for a in data["animals"] if a.get("arete")}
    by_name = {a["name"].lower(): a["id"] for a in data["animals"] if a.get("name")}

    def resolve_animal_id(v):
        s = norm(v)
        if not s:
            return ""
        digits = re.sub(r"[^\d]", "", s)
        if digits and digits in by_arete:
            return by_arete[digits]
        if s in by_arete:
            return by_arete[s]
        low = s.lower()
        if low in by_name:
            return by_name[low]
        return ""

    # --- Brutos from Bovinos_Bruto (con extras)
    if "Bovinos_Bruto" in wb.sheetnames:
        ws = wb["Bovinos_Bruto"]
        header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        mp = find_header_map(header)
        i = 0

        for r in range(2, ws.max_row + 1):
            row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            if all(v is None or str(v).strip() == "" for v in row):
                continue

            i += 1
            nombreArete = safe_get(row, mp, "nombre/arete", "nombre", "arete")
            estado = safe_get(row, mp, "estado", "estado_nota", "nota", "observacion")
            edad = safe_get(row, mp, "edad", "edad_meses", "meses")
            peso = safe_get(row, mp, "peso", "peso_kg", "kg")

            extras = {}
            for j, h in enumerate(header):
                k = slug_key(h)
                if not k:
                    continue
                v = row[j] if j < len(row) else None
                if v is None or str(v).strip() == "":
                    continue
                extras[k] = v

            data["brutos"].append({
                "id": uid("bru", i),
                "nombreArete": norm(nombreArete),
                "estadoNota": norm(estado),
                "edad": norm(edad),
                "peso": norm(peso),
                "extras": extras,
                "createdBy": "user_import",
                "createdAt": 0
            })

    # --- Milk from Ordeño (si existe)
    if "Ordeño" in wb.sheetnames:
        ws = wb["Ordeño"]
        header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        mp = find_header_map(header)
        idx = 0

        for r in range(2, ws.max_row + 1):
            row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            fecha = safe_get(row, mp, "fecha")
            arete = safe_get(row, mp, "arete")
            nombre = safe_get(row, mp, "nombre", "nombre/arete")
            lm = safe_get(row, mp, "litros mañana", "litros manana", "mañana", "manana")
            lt = safe_get(row, mp, "litros tarde", "tarde")

            if fecha is None and (lm is None and lt is None):
                continue
            date_iso = parse_date(fecha)
            if not date_iso:
                continue

            animal_id = resolve_animal_id(arete) or resolve_animal_id(nombre)
            if not animal_id:
                continue

            try:
                m = float(lm or 0)
            except Exception:
                m = 0.0
            try:
                t = float(lt or 0)
            except Exception:
                t = 0.0

            idx += 1
            data["milk"].append({
                "id": uid("milk", idx),
                "date": date_iso,
                "animalId": animal_id,
                "m": m,
                "t": t,
                "total": m + t,
                "createdBy": "user_import",
                "createdAt": 0
            })

    # --- Medicamentos -> meds + healthEvents + boosters
    if "Medicamentos" in wb.sheetnames:
        ws = wb["Medicamentos"]
        header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        mp = find_header_map(header)

        ev_i = 0
        boo_i = 0

        for r in range(2, ws.max_row + 1):
            row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]

            finca = safe_get(row, mp, "finca")
            arete = safe_get(row, mp, "arete", "id", "codigo")
            nombre = safe_get(row, mp, "nombre", "nombre/arete")

            fecha = safe_get(row, mp, "fecha", "fecha_aplicacion", "dia")
            proc = safe_get(row, mp, "medicamento/procedimiento", "procedimiento", "medicamento", "tratamiento")
            plan = safe_get(row, mp, "plan", "dosis", "detalle")
            responsable = safe_get(row, mp, "responsable", "veterinario", "aplico")
            costo = safe_get(row, mp, "costo", "valor", "precio")
            notas = safe_get(row, mp, "notas", "observacion", "obs")

            if (arete is None and nombre is None) and (proc is None and plan is None and fecha is None):
                continue

            date_iso = parse_date(fecha) or ""
            animal_id = resolve_animal_id(arete) or resolve_animal_id(nombre) or resolve_animal_id(str(nombre))
            if not animal_id:
                animal_id = ""  # lo guardamos igual en meds

            # ✅ meds
            ev_i += 1
            data["meds"].append({
                "id": uid("med", ev_i),
                "animalId": animal_id,
                "nombre": norm(nombre),
                "fecha": date_iso,
                "procedimiento": norm(proc),
                "responsable": norm(responsable),
                "plan": norm(plan),
                "costo": norm(costo),
                "notas": norm(notas),
                "finca": norm(finca),
                "createdBy": "user_import",
                "createdAt": 0
            })

            # ✅ también como healthEvents (si hay animal_id)
            if animal_id:
                event_id = uid("hev", ev_i)
                data["healthEvents"].append({
                    "id": event_id,
                    "animalId": animal_id,
                    "procedure": norm(proc) or norm(plan) or "Medicamento",
                    "date": date_iso,
                    "createdBy": "user_import",
                    "createdAt": 0
                })

                text = f"{norm(plan)} {norm(proc)}"
                for mm in re.finditer(r"(\d{1,2})[-/](\d{1,2})[-/](\d{2,4})", text):
                    d, mo, y = int(mm.group(1)), int(mm.group(2)), int(mm.group(3))
                    if y < 100:
                        y += 2000
                    try:
                        ref = datetime.date(y, mo, d).isoformat()
                    except ValueError:
                        continue
                    boo_i += 1
                    data["boosters"].append({
                        "id": uid("boo", boo_i),
                        "eventId": event_id,
                        "animalId": animal_id,
                        "procedure": norm(proc) or "Refuerzo",
                        "refDate": ref,
                        "finca": norm(finca),
                        "status": "pending",
                        "createdBy": "user_import",
                        "createdAt": 0
                    })

    # --- Repro from Reproduccion (si existe)
    if "Reproduccion" in wb.sheetnames:
        ws = wb["Reproduccion"]
        header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        mp = find_header_map(header)
        rep_i = 0

        for r in range(2, ws.max_row + 1):
            row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            arete = safe_get(row, mp, "arete")
            nombre = safe_get(row, mp, "nombre", "nombre/arete")
            parto = safe_get(row, mp, "fecha último parto", "fecha ultimo parto", "ultimo parto", "parto")
            celo = safe_get(row, mp, "fecha último celo", "fecha ultimo celo", "ultimo celo", "celo")
            insem = safe_get(row, mp, "fecha inseminación", "fecha inseminacion", "inseminacion")
            pre = safe_get(row, mp, "diagnóstico preñez (si/no)", "diagnostico preñez (si/no)", "preñez", "prenhez")

            if arete is None and nombre is None and parto is None and celo is None and insem is None and pre is None:
                continue

            animal_id = resolve_animal_id(arete) or resolve_animal_id(nombre)
            if not animal_id:
                continue

            rep_i += 1
            data["repro"].append({
                "id": uid("rep", rep_i),
                "animalId": animal_id,
                "parto": parse_date(parto),
                "celo": parse_date(celo),
                "insem": parse_date(insem),
                "pre": norm(pre).upper(),
                "createdBy": "user_import",
                "createdAt": 0
            })

    # --- Finance helpers
    def import_simple(sheet, mapping):
        if sheet not in wb.sheetnames:
            return []
        ws = wb[sheet]
        header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        mp = find_header_map(header)
        out = []
        for r in range(2, ws.max_row + 1):
            row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            if all(v is None or str(v).strip() == "" for v in row):
                continue
            rec = {}
            for out_key, in_names in mapping.items():
                rec[out_key] = safe_get(row, mp, *in_names)
            out.append(rec)
        return out

    # Venta_Queso
    for i, rec in enumerate(import_simple("Venta_Queso", {
        "date": ["fecha"],
        "client": ["cliente"],
        "lbs": ["libras"],
        "price": ["precio (cop)", "precio"],
        "total": ["total (cop)", "total"]
    }), start=1):
        date_iso = parse_date(rec["date"])
        if not date_iso:
            continue
        lbs = float(rec["lbs"] or 0)
        price = float(rec["price"] or 0)
        total = float(rec["total"] or (lbs * price))
        data["salesCheese"].append({
            "id": uid("sale", i),
            "date": date_iso,
            "client": norm(rec["client"]),
            "lbs": lbs,
            "price": price,
            "total": total,
            "createdBy": "user_import",
            "createdAt": 0
        })

    # Compra_Leche
    for i, rec in enumerate(import_simple("Compra_Leche", {
        "period": ["periodo"],
        "liters": ["litros"],
        "vl": ["valor/litro (cop)", "valor/litro"],
        "total": ["total (cop)", "total"]
    }), start=1):
        if not norm(rec["period"]) and rec["liters"] is None:
            continue
        liters = float(rec["liters"] or 0)
        vl = float(rec["vl"] or 0)
        total = float(rec["total"] or (liters * vl))
        data["buyMilk"].append({
            "id": uid("buy", i),
            "period": norm(rec["period"]),
            "liters": liters,
            "vl": vl,
            "total": total,
            "createdBy": "user_import",
            "createdAt": 0
        })

    # Transporte_Leche
    for i, rec in enumerate(import_simple("Transporte_Leche", {
        "period": ["periodo"],
        "value": ["valor transporte (cop)", "valor transporte"],
        "qty": ["cantidad"],
        "total": ["total (cop)", "total"]
    }), start=1):
        if not norm(rec["period"]) and rec["value"] is None:
            continue
        value = float(rec["value"] or 0)
        qty = float(rec["qty"] or 0)
        total = float(rec["total"] or (value * qty))
        data["transMilk"].append({
            "id": uid("tm", i),
            "period": norm(rec["period"]),
            "value": value,
            "qty": qty,
            "total": total,
            "createdBy": "user_import",
            "createdAt": 0
        })

    # Gastos_Fijos
    for i, rec in enumerate(import_simple("Gastos_Fijos", {
        "concept": ["concepto"],
        "value": ["valor mensual (cop)", "valor mensual"],
        "notes": ["notas"]
    }), start=1):
        if not norm(rec["concept"]) and rec["value"] is None:
            continue
        value = float(rec["value"] or 0)
        data["fixedCosts"].append({
            "id": uid("fx", i),
            "concept": norm(rec["concept"]),
            "value": value,
            "createdBy": "user_import",
            "createdAt": 0
        })

    out = {
        "exportedAt": datetime.datetime.utcnow().isoformat() + "Z",
        "data": data
    }
    Path("ganaderia_import.json").write_text(
        json.dumps(out, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )
    print("OK -> ganaderia_import.json creado")


if __name__ == "__main__":
    main()
        data["fixedCosts"].append({"id":uid("fx", idx),"concept":norm(rec["concept"]),
                                  "value":value,"createdBy":"user_import","createdAt":0})

    out = {"exportedAt": datetime.datetime.utcnow().isoformat()+"Z", "data": data}
    Path("ganaderia_import.json").write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    print("OK -> ganaderia_import.json creado")

if __name__ == "__main__":
    main()
